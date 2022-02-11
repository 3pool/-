# -*- coding:utf-8 -*-
# 抓取51job数据分析职位详情页内容，并写入本地文件
'''
1.抓目录页内容，把详情页链接保存到本地文件
1.1 抓目录页内容：selenium
1.2 抓链接信息：bs/re
1.3 把抓取到的信息写入本地文件：openpyxl
2.抓取详情页内容
2.1 打开保存的链接文件，获取详情页链接
2.2 针对详情页进行信息抓取：selenium + re
3.把详情页内容写入磁盘文件：openpyxl
'''

# 0.全局处理
# 载入库
from selenium import webdriver
# 载入webdriver对象
from selenium.webdriver.chrome.options import Options
# 载入Option，方便后续用无头模式
import openpyxl
import re
from bs4 import BeautifulSoup as bs
import urllib.request
import requests
import time

# 全局变量
menuList = []
# 存储目录页信息
jobList = []
# 存储详情页信息

# 测试，用requests抓取
def get_links_by_requests():
    url = 'https://search.51job.com/list/010000%252C030200%252C040000%252C020000%252C01,000000,0000,00,9,99,%25E6%2595%25B0%25E6%258D%25AE%25E5%2588%2586%25E6%259E%2590,2,1.html?lang=c&postchannel=0000&workyear=99&cotype=99&degreefrom=99&jobterm=99&companysize=99&ord_field=0&dibiaoid=0&line=&welfare='
    response = requests.get(url,headers = {'User-Agent':'Mozilla / 5.0'})
    # print(response.text)    # 调试用
    # 结果证明51job的搜索结果是通过js动态生成的，用requests和urllib抓取会很麻烦

# 1.抓目录页内容，把详情页链接保存到本地文件
# 1.1 抓目录页内容：selenium
def get_job_links(pageno):
    # pageno是要抓取的目录页的数量
    # 1.1.1 构建一个浏览器对象
    browser = webdriver.Chrome()
    url = 'http://www.51job.com'
    browser.set_window_size(600,800)
    # 设置浏览器窗口的大小
    # 1.1.2 开启浏览器
    browser.get(url)
    time.sleep(5)
    # 暂停5秒
    # 1.1.3 控制浏览器进行搜索
    # 通过ID搜索输入框
    inputBox = browser.find_element_by_id('kwdselectid')
    # 在输入框中输入信息
    inputBox.send_keys('数据分析')
    time.sleep(3)
    # 通过xpath寻找搜索按钮
    submitButton = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div/button')
    # 模拟鼠标点击
    submitButton.click()
    time.sleep(3)
    # input('敲回车继续')

    # 1.2 获取目录页内容
    for i in range(pageno):
        # 设置翻页的循环
        # 1.2.1 抓取当前页面的内容
        response = browser.page_source
        # print(response) # 调试用
        # 转成bs对象
        jobSoup = bs(response,'html.parser')
        tagJobDiv = jobSoup.find('div', class_='j_joblist')
        # print(tagJobDiv)    # 调试用
        # 搜索职位名称、薪酬、公司、工作地点等、详情页链接
        # 职位名称
        tagTitle = tagJobDiv.find_all('span', class_='jname at')
        # 返回结果是包含所有职位名称的span标签的列表
        # print(tagTitle)   # 调试用
        # print(len(tagTitle))  # 调试用
        # 薪酬
        tagSalary = tagJobDiv.find_all('span', class_='sal')
        # print(tagSalary)    # 调试用
        # 工作地点
        tagAt = tagJobDiv.find_all('span', class_='d at')
        # 公司
        tagCom = tagJobDiv.find_all('a', class_='cname at')
        # 详情页链接
        tagLink = tagJobDiv.find_all('a', class_='el')

        # 提取所有抓取到的列表中的内容
        for i in range(len(tagTitle)):
            # 用索引同时遍历5个结果集
            print('职位名称：',tagTitle[i]['title'])
            print('薪酬：',tagSalary[i].text)
            print('工作地点等：',tagAt[i].text)
            print('公司：',tagCom[i]['title'])
            print('详情页链接：',tagLink[i]['href'])

            print('-' * 100)
            menuList.append((tagTitle[i]['title'],tagSalary[i].text,tagAt[i].text,tagCom[i]['title'],tagLink[i]['href']))
        # input()

        # 搜索翻页的链接
        nextPage = browser.find_element_by_xpath('/html/body/div[2]/div[3]/div/div[2]/div[4]/div[2]/div/div/div/ul/li[8]/a')
        # 点击翻页链接
        nextPage.click()
        time.sleep(5)

# 1.3 把抓取到的信息写入本地文件：openpyxl
def save_menu_info():
    iBook = openpyxl.Workbook()
    # 构建工作簿
    iSheet = iBook.create_sheet('51job目录',0)
    # 创建工作表

    # 写表头
    iSheet.cell(1, 1).value = '编号'
    iSheet.cell(1, 2).value = '职位名称'
    iSheet.cell(1, 3).value = '薪酬'
    iSheet.cell(1, 4).value = '工作地点等'
    iSheet.cell(1, 5).value = '公司'
    iSheet.cell(1, 6).value = '详情页链接'

    # 写数据
    for iRow in range(len(menuList)):
        # 外层循环处理行，行数就是menuList的元素数量
        iSheet.cell(iRow+2,1).value = iRow+1
        # 在外层循环中写入编号
        for iCol in range(len(menuList[iRow])):
            # 内层循环处理列，列数就是menuList中每个职位的字段数量
            iSheet.cell(iRow+2,iCol+2).value = menuList[iRow][iCol]
            # 在指定位置写入对应的内容

    # 保存工作簿到本地磁盘文件
    iBook.save('51job目录.xlsx')

# 2.抓取详情页内容
# 2.1 打开保存的链接文件，获取详情页链接
def get_links():
    jBook = openpyxl.load_workbook('51job目录.xlsx')
    # 从磁盘文件读取工作簿
    jSheet = jBook.worksheets[0]
    # 载入工作簿的第一张表，即目录表
    # print(jSheet) # 调试用
    # print(jSheet.columns) # 调试用
    # jSheet.columns返回所有列，但结果是一个生成器，不便进行后续处理，可以转换为元组或列表进行后续操作
    jColumns = tuple(jSheet.columns)
    # 返回的结果是由每列的cell对象的元组构成的嵌套的元组
    # print(jColumns) # 调试用
    # print(jColumns[5]) # 调试用
    # 返回第6列，即详情页链接所在的列
    jobLinks = []
    # 构建一个列表，用来存储详情页链接
    for jCell in jColumns[5][1:]:
        # 遍历第六列（链接列）除表头之外的所有cell
        # print(jCell.value)  # t调试用
        jobLinks.append(jCell.value)
        # 将每个cell中的链接append到jobLinks中
    # print(jobLinks)  # 调试用
    # print(len(jobLinks))    # 调试用
    return jobLinks
    # 将获取到的jobLinks返回给调用该函数的对象

# 2.2 针对详情页进行信息抓取：selenium + re
def get_full_data(urls):
    # 调用时，直接把get_links()作为参数放到函数的（）里
    # 因为selenium速度很慢，所以测试的时候先用一两个职位进行测试，测试完成后再针对所有的页面进行抓取
    for jobLink in urls[:3]:
        print(jobLink)  # 调试用
        data_page = ''
        try:
            # 构建无头浏览器进行抓取，以提高速度，减少资源的占用
            # ch_options = Options()
            # 构建一个Options对象
            # ch_options.add_argument('--headless')
            # 给Option对象添加一个无头的参数
            browser = webdriver.Chrome()
            # browser = webdriver.Chrome(options=ch_options)
            # 构建一个浏览器，同时设定对应的options
            browser.get(jobLink)
            # 访问详情页
            time.sleep(5)
            data_page = browser.page_source
            browser.quit()
            # 关闭浏览器
        except:
            pass
        # print(data_page)    # 调试用
        # 2.2.1 职位名称
        pn_jobtitle = re.compile(r'<div class="tHeader tHjob">.*?<h1 title="(.*?)">',re.S)
        # 职位名称的正则表达式
        jobTitle = ''
        # 先构建一个空字符串
        try:
            jobTitle = re.search(pn_jobtitle,data_page).group(1)
            # 在group中加入参数1，是返回用()分组的正则表达式中的内容
        except:
            jobTitle = ''
        print('职位名称：',jobTitle)

        # 2.2.2 薪酬
        pn_jobsalary = re.compile(r'<div class="tHeader tHjob">.*?<strong>(.*?)</strong>',re.S)
        jobSalary = ''
        try:
            jobSalary = re.findall(pn_jobsalary,data_page)[0]
            # 用索引返回结果集中的第一项
        except:
            jobSalary = ''
        print('薪酬：',jobSalary)

        # 2.2.3 公司名称
        jobSoup = bs(data_page,'html.parser')
        # 构建soup对象
        tagJobHeader = jobSoup.find('div',class_='tHeader tHjob')
        jobCompany = ''
        try:
            jobCompany = tagJobHeader.find('a',class_='catn')['title']
        except:
            jobCompany = ''
        print('公司名称：',jobCompany)

        # 2.2.4 工作地点，工作经验，学历
        pn_jobmsg = re.compile(r'<div class="tHeader tHjob">.*?<p class="msg ltype" title="(.*?)">',re.S)
        jobMsg = ''
        jobArea = ''
        jobExp = ''
        jobEdu = ''
        try:
            jobMsg = re.findall(pn_jobmsg,data_page)[0].split('&nbsp;&nbsp;|&nbsp;&nbsp;')
            jobArea = jobMsg[0]
            jobExp = jobMsg[1]
            jobEdu = jobMsg[2]
        except:
            pass
        print('工作地点：',jobArea)
        print('工作经验：',jobExp)
        print('学历：',jobEdu)

        # 2.2.5 福利待遇
        # 先抓包含全部福利的div，因为有些没有福利的职位，但div是有的
        jobWelfare = ''
        try:
            tagWelfareDiv = jobSoup.find('div',class_='jtag')
            tagWelfareSpan = tagWelfareDiv.find_all('span')
            if len(tagWelfareSpan) == 0:
                # 没有福利待遇
                jobWelfare = '无福利'
            else:
                for welfare in tagWelfareSpan:
                    jobWelfare += '/' + welfare.text
                jobWelfare = jobWelfare[1:]
        except:
            jobWelfare = '无福利'
        print('福利待遇：',jobWelfare)

        # 2.2.6 职位信息
        jobDesc = ''
        pn_jobdesc = re.compile(r'<div class="bmsg job_msg inbox">(.*?)<div class="mt10">',re.S)
        try:
            jobDesc = re.sub(r'<.*?>','',re.findall(pn_jobdesc,data_page)[0].strip())
            # 用findall返回所有的职位描述，用[0]取出第一项，用strip()去除两端的不可见字符，再用re.sub替换掉所有的html标签，只保留文本
        except:
            pass
        print('职位描述：',jobDesc)

        # 2.2.7 公司性质、规模、行业
        tagComDiv = jobSoup.find('div', class_='com_tag')
        comType = ''
        comScale = ''
        comIndu = ''
        try:
            comInfo = tagComDiv.find_all('p',class_='at')
            comType = comInfo[0]['title']
            comScale = comInfo[1]['title']
            comIndu = comInfo[2]['title']
        except:
            pass
        print('公司性质：',comType)
        print('公司规模：',comScale)
        print('行业类别：',comIndu)

        print('-'*100)

        # 将所有抓取到的信息append到joblist中
        if jobTitle != '':
            # 职位名称不为空时添加
            jobList.append((jobTitle,jobSalary,jobCompany,jobArea,jobExp,jobEdu,jobWelfare,jobDesc,comType,comScale,comIndu))

# 测试函数
# get_links_by_requests()
# get_job_links(3)     # 抓取目录页
# save_menu_info()    # 写入目录页
# get_links()   # 从本地文件读取链接
get_full_data(get_links())
print(jobList)
print(len(jobList))


'''
<span title="技术开发岗（数据）" class="jname at">技术开发岗（数据）</span>
'''
'''
搜索输入框
<input type="text" maxlength="200" id="kwdselectid" placeholder="请输入关键字" vindex="-1" value="" class="mytxt at" preval="%E6%95%B0%E6%8D%AE%E5%88%86%E6%9E%90">
'''
'''
单个职位
<div class="e"><input class="checkbox" jt="0_0" name="delivery_jobid" style="display: none;" type="checkbox" value="125688571"/> <input class="checkbox" jt="0_0" name="delivery_jobid_125688571" style="display: none;" type="checkbox" value="125688571"/> <div class="e_icons ick"></div> <a class="el" href="https://jobs.51job.com/nanjing-jyq/125688571.html?s=sou_sou_soulb&amp;t=0_0" target="_blank"><p class="t"><span class="jname at" title="数据分析师">数据分析师</span> <span class="time">07-24发布</span> <!-- --> <!-- --> <!-- --> <!-- --> <em alt="闪" class="e_icons tbs sd" title="HR会快速回复该职位的投递申请"></em></p> <p class="info"><span class="sal">5-9千/月</span> <span class="d at">南京-建邺区  |  1年经验  |  本科  |  招5人</span></p> <p class="tags" title="五险一金 弹性工作 交通补贴 餐饮补贴 通讯补贴 年终奖金 13薪"><span><!-- --> <i>五险一金</i><i>弹性工作</i><i>交通补贴</i><i>餐饮补贴</i><i>通讯补贴</i><i>年终奖金</i><i>13薪</i></span></p></a> <div class="er"><a class="cname at" href="https://jobs.51job.com/all/co6086507.html" target="_blank" title="通联数据股份公司">通联数据股份公司</a> <p class="dc at">合资 | 5000-10000人</p> <p class="int at">金融/投资/证券</p> <button class="p_but" event-type="99" trace-name="申请职位-职位下" track-type="searchTrackButtonClick">申请职位</button></div></div>

'''

'''
详情页
<div class="tHeader tHjob">
    <div class="in">
        <div class="cn">
                        <h1 title="数据分析师">数据分析师<input value="125688571" name="hidJobID" id="hidJobID" type="hidden" jt="0_0"><img title="HR会快速回复该职位的投递申请" src="//img02.51jobcdn.com/im/jobs/icon_lightning.png" width="18" height="18" alt=""></h1><strong>5-9千/月</strong>
            <p class="cname">
                <a href="https://jobs.51job.com/all/co6086507.html" target="_blank" title="通联数据股份公司" class="catn">通联数据股份公司<em class="icon_b i_link"></em></a>
                                    <a track-type="jobsButtonClick" event-type="2" class="i_house" href="https://jobs.51job.com/all/co6086507.html?#syzw" target="_blank">查看所有职位</a>
                            </p>
            <p class="msg ltype" title="南京-建邺区&nbsp;&nbsp;|&nbsp;&nbsp;1年经验&nbsp;&nbsp;|&nbsp;&nbsp;本科&nbsp;&nbsp;|&nbsp;&nbsp;招5人&nbsp;&nbsp;|&nbsp;&nbsp;07-24发布">南京-建邺区&nbsp;&nbsp;<span>|</span>&nbsp;&nbsp;1年经验&nbsp;&nbsp;<span>|</span>&nbsp;&nbsp;本科&nbsp;&nbsp;<span>|</span>&nbsp;&nbsp;招5人&nbsp;&nbsp;<span>|</span>&nbsp;&nbsp;07-24发布</p>
            <div class="jtag">
                <div class="t1">
                                            <span class="sp4">五险一金</span><span class="sp4">弹性工作</span><span class="sp4">交通补贴</span><span class="sp4">餐饮补贴</span><span class="sp4">通讯补贴</span><span class="sp4">年终奖金</span><span class="sp4">13薪</span>                                        <div class="clear"></div>
                </div>
            </div>
        </div>
        <div class="op">
            <a track-type="jobsButtonClick" event-type="1" class="but_sq" id="app_ck" href="javascript:void(0);" onclick="delivery('hidJobID', '1', '//i.51job.com', 'c', '', 'sou_sou_soulb', 'detail', '//img03.51jobcdn.com');return false;">	<img width="19" height="24" src="//img01.51jobcdn.com/im/jobs/but_sq_arr.png" alt="">申请职位</a>
            <a track-type="jobsButtonClick" event-type="4" class="icon_b i_upline" id="125688571" target="_blank" href="//i.51job.com/userset/bounce_window_redirect.php?jobid=125688571&amp;redirect_type=2">竞争力分析</a>
            <a track-type="jobsButtonClick" event-type="3" class="icon_b i_collect" href="javascript:void(0);" onclick="saveCollection('125688571');return false;">收藏</a>
            <div class="clear"></div>
        </div>
    </div>
</div>

'''

'''
详情页职位信息
<div class="tBorderTop_box">
    <h2 class="prop">
        <span class="bname">职位信息</span>
            </h2>
    <div class="bmsg job_msg inbox">
        <p>1、负责股票、债券、基金等金融数据日常加工生产，对数据进行及时跟进处理入库</p><p>2、对自己负责的数据的正确率及准确性负责，及时的添加校验，梳理数据逻辑</p><p>3、保证新增数据的及时入库，每月完成组内安排的修复任务</p><p>4、针对负责的数据表，完善加工规则，追踪数据源，对数据进行深度挖掘</p><p>5、不断优化数据生产方式，并创新数据处理的工具改进优化</p><p>6、针对行业情况进行分析，协助或撰写行业研究报告</p><p>7、及时与大客户沟通，了解客户需求，提供数据解决方案</p><p><br></p><p>任职要求：</p><p>1、本科以上学历，经济学、金融学、金融统计、数学、会计等相关专业毕业</p><p>2、1年以上工作经验</p><p>3、具有一定的金融基础知识，对数据处理有一定经验</p><p>4、对SQL语句了解，能熟练书写数据库查询语句</p><p>5、能够熟练使用OFFICE办公软件，对数字敏感</p><p>6、具有较强的思维逻辑，语言表达清晰善于沟通，反应灵活</p>
                <div class="mt10">
                                        <p class="fp"><span class="label">职能类别：</span><a class="el tdn" href="https://jobs.51job.com/nanjing-jyq/shujufenxishi/">数据分析师</a><a class="el tdn" href="https://jobs.51job.com/nanjing-jyq/shujuyunying/">数据运营</a></p>
                                        <p class="fp"><span class="label">关键字：</span><a class="el tdn" href="https://search.51job.com/list/070200,070204,0000,00,9,99,%25E6%2595%25B0%25E6%258D%25AE,2,1.html?lang=c&amp;postchannel=0000&amp;workyear=99&amp;cotype=99&amp;degreefrom=99&amp;jobterm=99&amp;companysize=99&amp;ord_field=0&amp;dibiaoid=0&amp;line=&amp;welfare=">数据</a><a class="el tdn" href="https://search.51job.com/list/070200,070204,0000,00,9,99,%25E5%2588%2586%25E6%259E%2590,2,1.html?lang=c&amp;postchannel=0000&amp;workyear=99&amp;cotype=99&amp;degreefrom=99&amp;jobterm=99&amp;companysize=99&amp;ord_field=0&amp;dibiaoid=0&amp;line=&amp;welfare=">分析</a><a class="el tdn" href="https://search.51job.com/list/070200,070204,0000,00,9,99,%25E6%2595%25B0%25E6%258D%25AE%25E6%258C%2596%25E6%258E%2598,2,1.html?lang=c&amp;postchannel=0000&amp;workyear=99&amp;cotype=99&amp;degreefrom=99&amp;jobterm=99&amp;companysize=99&amp;ord_field=0&amp;dibiaoid=0&amp;line=&amp;welfare=">数据挖掘</a><a class="el tdn" href="https://search.51job.com/list/070200,070204,0000,00,9,99,%25E7%25BB%259F%25E8%25AE%25A1,2,1.html?lang=c&amp;postchannel=0000&amp;workyear=99&amp;cotype=99&amp;degreefrom=99&amp;jobterm=99&amp;companysize=99&amp;ord_field=0&amp;dibiaoid=0&amp;line=&amp;welfare=">统计</a><a class="el tdn" href="https://search.51job.com/list/070200,070204,0000,00,9,99,%25E6%2595%25B0%25E6%258D%25AE%25E5%2588%2586%25E6%259E%2590%25E5%25B8%2588,2,1.html?lang=c&amp;postchannel=0000&amp;workyear=99&amp;cotype=99&amp;degreefrom=99&amp;jobterm=99&amp;companysize=99&amp;ord_field=0&amp;dibiaoid=0&amp;line=&amp;welfare=">数据分析师</a><a class="el tdn" href="https://search.51job.com/list/070200,070204,0000,00,9,99,sql,2,1.html?lang=c&amp;postchannel=0000&amp;workyear=99&amp;cotype=99&amp;degreefrom=99&amp;jobterm=99&amp;companysize=99&amp;ord_field=0&amp;dibiaoid=0&amp;line=&amp;welfare=">sql</a></p>
                                </div>
        <div class="share"><a track-type="jobsButtonClick" event-type="6" class="a" href="javascript:void(0);" onclick="weixinMa();">微信分享</a><div id="weixinMa_fx" style="display:none;"><img width="198" height="198" alt="二维码" org="https://jobs.51job.com/wx_qrcode.php?url=https%3A%2F%2Fm.51job.com%2Fsearch%2Fjobdetail.php%3Fjobid%3D125688571"></div>
        </div>
        <div class="clear"></div>
    </div>
</div>
'''

'''

        <p>1、负责股票、债券、基金等金融数据日常加工生产，对数据进行及时跟进处理入库</p><p>2、对自己负责的数据的正确率及准确性负责，及时的添加校验，梳理数据逻辑</p><p>3、保证新增数据的及时入库，每月完成组内安排的修复任务</p><p>4、针对负责的数据表，完善加工规则，追踪数据源，对数据进行深度挖掘</p><p>5、不断优化数据生产方式，并创新数据处理的工具改进优化</p><p>6、针对行业情况进行分析，协助或撰写行业研究报告</p><p>7、及时与大客户沟通，了解客户需求，提供数据解决方案</p><p><br></p><p>任职要求：</p><p>1、本科以上学历，经济学、金融学、金融统计、数学、会计等相关专业毕业</p><p>2、1年以上工作经验</p><p>3、具有一定的金融基础知识，对数据处理有一定经验</p><p>4、对SQL语句了解，能熟练书写数据库查询语句</p><p>5、能够熟练使用OFFICE办公软件，对数字敏感</p><p>6、具有较强的思维逻辑，语言表达清晰善于沟通，反应灵活</p>
                


'''

'''
公司情况
<div class="com_tag">
            <p class="at" title="合资"><span class="i_flag"></span>合资</p>
            <p class="at" title="5000-10000人"><span class="i_people"></span>5000-10000人</p>
            <p class="at" title="金融/投资/证券">
                <span class="i_trade"></span>
                                    <a href="https://jobs.51job.com/nanjing/hy03/">金融/投资/证券</a>
                            </p>
        </div>
'''