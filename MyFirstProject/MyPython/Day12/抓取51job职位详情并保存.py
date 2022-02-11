# -*- coding:utf-8 -*-
# 抓取51job数据分析职位详情页内容，并写入本地文件
'''
1.抓目录页内容，把详情页链接保存到本地文件
1.1 抓目录页内容：selenium
1.2 抓链接信息：bs/re
1.3 把抓取到的信息写入本地文件：openpyxl
2.抓取详情页内容
2.1 打开保存的链接文件，获取详情页链接
2.2 针对详情页进行信息抓取：selenium + re
3.把详情页内容写入磁盘文件：openpyxl
'''
import os
from selenium import webdriver
from bs4 import BeautifulSoup as bs
import re
import openpyxl
import time
from selenium.webdriver.chrome.options import Options

# 定义全局变量
# 创建列表用于存储详情页链接
menuList = []
# 创建列表用于存储详情页信息
jobList = []
# 获取目录页信息
def get_joblinks(pageno):
    # 创建一个浏览器对象
    browser = webdriver.Chrome()
    url = 'http://www.51job.com'
    # 设置浏览器窗口大小
    browser.set_window_size(600,800)
    # 打开浏览器
    browser.get(url)
    # 暂停5秒
    time.sleep(5)
    # 通过ID查找搜索框
    inputBox = browser.find_element_by_id('kwdselectid')
    # 在搜索框中输入搜索内容
    inputBox.send_keys('数据分析')
    # 暂停3秒
    time.sleep(3)
    # 通过xpath查找搜索按钮
    submitButton = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div/button')
    # 模拟鼠标点击
    submitButton.click()
    # 暂停3秒
    time.sleep(3)
    # 创建循环，获取所有职位的URL
    for i in range(pageno):
        # 抓取当前页面的内容
        response = browser.page_source
        # print(response)
        # 转换成soup对象
        jobSoup = bs(response,'html.parser')
        # 获取包含所有信息的div
        tagJobDiv = jobSoup.find('div',class_ = 'j_joblist')
        # print(tagJobDiv)
        # 获取包含所有职位名称的列表
        tagTitles = tagJobDiv.find_all('span',class_ = 'jname at')
        # print(tagTitles)
        # print(len(tagTitles))
        # 获取包含所有职位薪酬的列表
        tagSalary = tagJobDiv.find_all('span',class_ = 'sal')
        # 获取包含所有公司名称的列表
        tagCompany = tagJobDiv.find_all('a',class_ = 'cname at')
        # 获取包含所有工作地点等信息的列表
        tagAt = tagJobDiv.find_all('span', class_='d at')
        # 获取包含所有职位详情页链接的列表
        tagLink = tagJobDiv.find_all('a', class_='el')
        # 提取所有抓到的列表中的内容
        for j in range(len(tagTitles)):
            print('职位名称：',tagTitles[j]['title'])
            print('职位薪酬：', tagSalary[j].text)
            print('公司名称：', tagCompany[j]['title'])
            print('工作地点等：', tagAt[j].text)
            print('详情页链接：', tagLink[j]['href'])
            print('-'*100)
            # 将抓到的内容存放到menulist中
            menuList.append((tagTitles[j]['title'],tagSalary[j].text,tagCompany[j]['title'],tagAt[j].text,tagLink[j]['href']))
        # 获取翻页的按钮
        nextPage = browser.find_element_by_css_selector('li.next>a')
        # 点击翻页链接
        nextPage.click()
        # 暂停5秒
        time.sleep(5)
        print('已完成第%d目录页信息的获取'%(i+1))
    print('已完成所有目录页信息的获取，共获取到(%d)个页面信息'%(i+1))

# 利用openpyxl把抓取到的信息写入本地的excel文件中
def save_menuinfo():
    # 创建工作簿
    iBook = openpyxl.Workbook()
    # 新建工作表
    iSheet = iBook.create_sheet('51job目录',0)
    # 写入表头
    iSheet.cell(1, 1).value = '编号'
    iSheet.cell(1, 2).value = '职位名称'
    iSheet.cell(1, 3).value = '薪酬'
    iSheet.cell(1, 4).value = '公司名称'
    iSheet.cell(1, 5).value = '工作地点等'
    iSheet.cell(1, 6).value = '详情页链接'
    # 写入预先存储在menulist中的职位信息
    # 外层循环处理行，行数就是menulist的元组数量
    for iRow in range(len(menuList)):
        # 外层循环写入编号
        iSheet.cell(iRow+2,1).value = iRow+1
        # 内层循环处理列，列数就是menulist每个元组的字段数量
        for iCol in range(len(menuList[iRow])):
            # 内层循环写入职位具体信息
            iSheet.cell(iRow+2,iCol+2).value = menuList[iRow][iCol]
    # 保存文件
    iBook.save('51job目录.xlsx')
    print('已完成目录页信息的保存')

# 抓取详情页内容
# 打开保存的目录文件，获取编号、职位名称及详情页链接
def get_links():
    # 打开工作簿
    jBook = openpyxl.load_workbook('51job目录.xlsx')
    # 获取工作表
    jSheet = jBook.worksheets[0]
    # print(jSheet)
    # print(jSheet.columns)
    # jSheet.columns返回所有列，但结果是一个生成器，不便进行后续处理，可以转换为元组或列表进行后续操作
    jColumns = tuple(jSheet.columns)
    # 返回的结果是由每列的cell对象的元组构成的嵌套的元组
    # 返回第6列，即详情页链接所在的列
    # print(jColumns)
    # print(len(jColumns))
    # 构建一个列表，用来存储详情页链接
    jobLinks = []
    for jCell6 in jColumns[5][2994:]:
        # 遍历第六列（链接列）除表头之外的所有cell
        # print(jCell.value)
        # 将每个cell中的链接append到jobLinks中
        jobLinks.append(jCell6.value)
        # print(jobLinks)
        # print(len(jobLinks))
    # 将获取到的jobLinks返回给调用该函数的对象
    jobNos = []
    for jCell1 in jColumns[0][2994:]:
        jobNos.append(jCell1.value)
    jobNames = []
    for jCell2 in jColumns[1][2994:]:
        jobNames.append(jCell2.value.replace('/','斜杠').replace('\\','反斜杠').replace('*','星号'))
    return jobNos,jobNames,jobLinks

# 通过抓取到的详情页链接，将每一页的详情页内容单独生成一个文件并保存到本地文件夹中
def save_linksinfo():
    os.makedirs('linksinfotmp',exist_ok=True)
    jobNos, jobNames, jobLinks = get_links()
    for i in jobNos:
        # 构建无头浏览器进行抓取，以提高速度，减少资源的占用
        # 构建一个option对象
        ch_options = Options()
        # 给Option对象添加一个无头的参数
        ch_options.add_argument('--headless')
        # 构建一个浏览器，同时设定对应的options
        browser = webdriver.Chrome(options=ch_options)
        # 访问详情页
        browser.get(jobLinks[i-2994])
        print('正在加载第%d页详情页内容到本地...'%(i))
        # 暂停5秒
        time.sleep(5)
        data_page = browser.page_source
        filename = 'linksinfotmp/' + str(i) + '_' + str(jobNames[i-2994]) + '.txt'
        with open(filename,'w',encoding='utf-8') as lif:
            lif.write(data_page)

# 针对详情页进行信息抓取：selenium + re
def get_fulldata():
    # 获取所有要抓取内容的职位文件名：用os.walk遍历指定的文件夹
    for root,dirs,files in os.walk('linksinfotmp'):
        # 获取文件夹对应的根目录、子目录和目录中的文件名，其中files对应的目录中的文件名列表
        pass
    # print(files)    #  调试用
    # print(len(files))   # 调试用
    for jobFile in files:
        # 遍历职位文件的列表
        with open('linksinfotmp/'+jobFile,'r',encoding='utf-8') as jf:
            data_page = jf.read()

        # 利用正则获取职位名称
        pn_jobTitle = re.compile(r'<div class="cn">.*?<h1 title="(.*?)">',re.S)
        # 创建空字符串用于存储职位名称
        jobTitle = ''
        try:
            # 获取职位名称
            jobTitle = re.search(pn_jobTitle,data_page).group(1)
            # 在group中加入参数1，是返回用()分组的正则表达式中的内容
        except:
            jobTitle = ''
        print('职位名称：',jobTitle)

        # 获取职位薪酬
        pn_jobSalary = re.compile(r'<div class="cn">.*?<strong>(.*?)</strong>',re.S)
        jobSalary = ''
        try:
            jobSalary = re.search(pn_jobSalary,data_page).group(1)
        except:
            jobSalary = ''
        print('职位薪酬：',jobSalary)

        # 获取公司名称
        # 创建soup对象
        jobSoup = bs(data_page,'html.parser')
        jobCompany = ''
        # 通过a标签的class属性查找公司名称
        try:
            jobCompany = jobSoup.find('a',class_ = 'catn')['title']
        except:
            jobCompany = ''
        print('公司名称：',jobCompany)

        # 获取工作地点、工作经验、学历等
        jobMsg = ''
        jobArea = ''
        jobExp = ''
        jobEdu = ''
        try:
            jobMsg = jobSoup.find('p',class_ = 'msg ltype')['title'].split('  |  ')
            jobArea = jobMsg[0]
            jobExp = jobMsg[1]
            jobEdu = jobMsg[2]
        except:
            pass
        # print(jobMsg)
        print('工作地点：',jobArea)
        print('工作经验：', jobExp)
        print('学历:', jobEdu)

        # 获取公司福利
        jobWelfare = ''
        try:
            tagWelfareDiv = jobSoup.find('div',class_ = 'jtag')
            tagWelfareSpan = tagWelfareDiv.find_all('span')
            if len(tagWelfareSpan) == 0:
                jobWelfare = '无福利'
            else:
                for welfare in tagWelfareSpan:
                    jobWelfare += '/' + welfare.text
                jobWelfare = jobWelfare[1:]
        except:
            jobWelfare = '无福利'
        print('福利待遇',jobWelfare)

        # 获取职位描述
        jobDesc = ''
        pn_jobdesc = re.compile(r'<div class="bmsg job_msg inbox">(.*?)<div class="mt10">',re.S)
        try:
            jobDesc = re.sub(r'<.*?>', '', re.findall(pn_jobdesc,data_page)[0].strip())
        except:
            pass
        print('职位描述：',jobDesc)

        # 获取公司性质、规模及行业
        tagComDiv = jobSoup.find('div',class_ = 'com_tag')
        comType = ''
        comScale = ''
        comIndu = ''
        try:
            comInfo = tagComDiv.find_all('p',class_ = 'at')
            comType = comInfo[0]['title']
            comScale = comInfo[1]['title']
            comIndu = comInfo[2]['title']
        except:
            pass
        print('公司性质：',comType)
        print('公司规模：', comScale)
        print('公司行业：', comIndu)

        print('-' * 100)

        # 将所有抓取到的信息写入到jobList中
        if jobTitle != '':
            # 职位名称不为空时添加
            jobList.append((jobTitle, jobSalary, jobCompany, jobArea, jobExp, jobEdu, jobWelfare, jobDesc, comType,comScale, comIndu))

# 将读取到的职位详情页信息保存到xlsx文件中
def sava_jobinfo():
    kBook = openpyxl.Workbook()
    kSheet = kBook.create_sheet('51job数据分析职位详情',0)
    kSheet.cell(1,1).value = '编号'
    kSheet.cell(1, 2).value = '职位名称'
    kSheet.cell(1, 3).value = '薪酬'
    kSheet.cell(1, 4).value = '公司名称'
    kSheet.cell(1, 5).value = '工作地点'
    kSheet.cell(1, 6).value = '工作经验'
    kSheet.cell(1, 7).value = '学历'
    kSheet.cell(1, 8).value = '福利待遇'
    kSheet.cell(1, 9).value = '职位描述'
    kSheet.cell(1, 10).value = '公司性质'
    kSheet.cell(1, 11).value = '公司规模'
    kSheet.cell(1, 12).value = '行业'
    for iRow in range(len(jobList)):
        kSheet.cell(iRow+2,1).value = iRow+1
        for iCol in range(len(jobList[iRow])):
            kSheet.cell(iRow+2,iCol+2).value = jobList[iRow][iCol]
    kBook.save('51job数据分析职位详情.xlsx')


# get_joblinks(80)
# save_menuinfo()
# get_links()
# save_linksinfo()
get_fulldata()
sava_jobinfo()


