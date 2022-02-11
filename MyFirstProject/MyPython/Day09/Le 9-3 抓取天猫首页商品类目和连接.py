# -*- coding:utf-8 -*-
'''
抓取天猫首页商品类目名称和链接，并写入到excel文件
思路：
1.抓取天猫首页页面内容：urllib
2.从抓取到的页面内容中提取商品类目和对应的链接：re
3.将抓取到的信息写入本地excel文件：openpyxl
'''
import urllib.request
import re
from openpyxl import Workbook

# 1.抓取天猫首页内容
# 1.1 先用简单粗暴的方法撩一下
url = 'https://www.tmall.com/'
response = urllib.request.urlopen(url)
strHtml = response.read().decode('utf-8')
# print(strHtml)  # 调试用

'''
                <a href="//nvzhuang.tmall.com/?acm=lb-zebra-148799-667863.1003.4.708026&amp;scm=1003.4.lb-zebra-148799-667863.OTHER_14561681423980_708026">女装</a>
                <a href="//neiyi.tmall.com/?acm=lb-zebra-148799-667863.1003.4.708026&amp;scm=1003.4.lb-zebra-148799-667863.OTHER_14561681423980_708026">内衣</a>
'''

# 2.从抓取到的页面内容中提取商品类目和对应的链接：re
# 2.1 分析要提取的信息
# 2.2 构建和要提取信息匹配的正则表达式
pnTmall = re.compile(r'<a href="//(.*?)">(.*?)</a>')
reTmall = re.findall(pnTmall,strHtml)
# print(reTmall)  # 调试用
# print(len(reTmall))  # 调试用
reTmall1 = reTmall[2:]
# 用切片的方法去掉不需要的信息
print(reTmall1)   # 调试用
print(len(reTmall1))  # 调试用
# for item in reTmall1:    # 调试用
#     print(item[1],item[0])

# 3.将抓取到的信息写入excel文件
tBook = Workbook()
# 创建一个新的工作簿
tSheet = tBook.create_sheet('天猫首页商品类目',0)
# 在索引为0的位置创建一个工作表
# 3.1 写表头
tSheet.cell(1,1).value = '编号'
tSheet.cell(1,2).value = '链接'
tSheet.cell(1,3).value = '类目名称'

# 3.2 写内容
for iRow in range(len(reTmall1)):
    # 外层循环处理行，行数由列表的元素数量决定
    # 写编号
    print(iRow)
    tSheet.cell(iRow+2,1).value = iRow + 1
    # 从第二行第一列开始写编号，因为python索引从0开始，第一行写入了表头，所以起始编号的行是iRow+2
    for iCol in range(len(reTmall1[iRow])):
        # 内层循环处理列，列数由列表中的元组的元素数量决定
        tSheet.cell(iRow+2,iCol+2).value = reTmall1[iRow][iCol]
        # 在指定的单元格中写入对应的数据
        # iRow+2,iCol+2是因为第一行写了表头，第一列写了编号，所以数据从第二行第二列开始写
        # reTmall1[iRow][iCol]是取得对应位置上的元素，方便写入
# tBook.save('tmall.xlsx')
# 将内存中的工作簿写入磁盘文件