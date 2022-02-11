# -*- coding:utf-8 -*-
# 抓取豆瓣电影top250首页的影片排名、影片名称、评分、评论人数和一句话短评
'''
思路:
1.用urllib抓取整个页面
2.用re提取需要的信息
3.用openpyxl写入文件
'''
# 载入库
import urllib.request
import re
from openpyxl import Workbook

# 1.用urllib抓取整个页面
def get_page():
    # 1.1 构建带有request headers的Request对象
    url = 'https://movie.douban.com/top250'
    myHeaders = {
        'Connection':'keep-alive',
        'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.90 Safari/537.36'
    }
    myRequest = urllib.request.Request(url,headers= myHeaders)
    # 1.2 开启Request对象，返回页面内容
    response = urllib.request.urlopen(myRequest)
    page = response.read().decode('utf-8')
    # print(page)     # 调试用
    # print(type(page))   # 调试用
    # 1.3 把抓取到的页面内容写入文本文件，方便后续的信息提取
    with open('top250.txt','w',encoding='utf-8') as file1:
        file1.write(page)
# 将抓取页面的步骤内容封装成函数方便调试

# 2.用re提取需要的信息
MovieList = []
# 构建一个空列表，存储所有影片的信息
def get_info():
    # 2.1 从磁盘文件读入页面内容
    with open('top250.txt','r',encoding='utf-8') as file2:
        page = file2.read()
    # print(page)   # 调试用
    # 2.2 提取包含所有影片的<li>标签
    pnLi = re.compile(r'<li>(.*?)</li>',re.S)
    reLi = re.findall(pnLi,page)
    # print('<li>:',reLi)     # 调试用
    # print(len(reLi))    # 调试用
    # 2.3 提取影片相关信息：排名、电影名称、评分、评论人数和短评
    for tagLi in reLi:
        # 遍历抓取到的所有的li标签，提取每部电影的信息
        # 2.3.1 提取影片排名
        pnNo = re.compile(r'<em class="">(.*?)</em>')
        MovieNo = re.findall(pnNo,tagLi)[0]
        # 在tagLi这个字符串中查找影片的排名
        print('影片排名：',MovieNo)
        # 2.3.2 提取影片名称
        pnName1 = re.compile(r'<span class="title">(.*?)</span>',re.S)
        MovieName1 = re.findall(pnName1,tagLi)[0]
        # print(MovieName1)
        pnName2 = re.compile(r'<img width="100" alt="(.*?)".*?>')
        MovieName2 = re.findall(pnName2,tagLi)[0]
        print('影片名称：',MovieName2)
        # 2.3.3 提取影片评分
        pnRating = re.compile(r'<span class="rating_num" property="v:average">(.*?)</span>',re.S)
        MovieRating = re.findall(pnRating,tagLi)[0]
        print('影片评分：',MovieRating)
        # 2.3.4 提取评论人数
        pnVote = re.compile(r'<span property="v:best" content="10.0"></span>.*?<span>(.*?)人评价</span>',re.S)
        MovieVote = re.findall(pnVote,tagLi)[0]
        print('评论人数：',MovieVote)
        # 2.3.5 提取一句话短评
        pnInq = re.compile(r'<span class="inq">(.*?)</span>',re.S)
        MovieInq = re.findall(pnInq,tagLi)[0]
        print('一句话短评：',MovieInq)
        print('-'*100)
        # 2.4 将所有抓取到的信息组合成嵌套的容器，构建一个结构化的数据，方便后续写入
        # 将每部影片的信息打包成元组，再添加到MovieList中，注意该操作是放在循环里的
        MovieList.append((MovieNo,MovieName2,MovieRating,MovieVote,MovieInq))

    # print('-'*100)  # 调试用
    # print(MovieList)    # 调试用
    # print(len(MovieList))   # 调试用

# 3.利用openpyxl写入excel文件
def sav2xlsx():
    myBook = Workbook()
    # 创建工作簿
    mySheet = myBook.create_sheet('豆瓣电影top2501',0)
    # 创建工作表

    # 写入表头
    mySheet.cell(1, 1).value = '影片排名'
    mySheet.cell(1, 2).value = '影片名称'
    mySheet.cell(1, 3).value = '影片评分'
    mySheet.cell(1, 4).value = '评论人数'
    mySheet.cell(1, 5).value = '一句话短评'

    # 写入数据
    for iRow in  range(len(MovieList)):
        # 外层循环处理行
        for iCol in range(len(MovieList[iRow])):
            # 内层循环处理列，列数就是列表中元组的元素数量
            mySheet.cell(iRow+2,iCol+1).value = MovieList[iRow][iCol]
            # 从第二行第一列开始写入数据，注意写入的位置和内容

    myBook.save('top2501.xlsx')

# 调试
get_info()
sav2xlsx()

'''
<li>
            <div class="item">
                <div class="pic">
                    <em class="">1</em>
                    <a href="https://movie.douban.com/subject/1292052/">
                        <img width="100" alt="肖申克的救赎" src="https://img2.doubanio.com/view/photo/s_ratio_poster/public/p480747492.jpg" class="">
                    </a>
                </div>
                <div class="info">
                    <div class="hd">
                        <a href="https://movie.douban.com/subject/1292052/" class="">
                            <span class="title">肖申克的救赎</span>
                                    <span class="title">&nbsp;/&nbsp;The Shawshank Redemption</span>
                                <span class="other">&nbsp;/&nbsp;月黑高飞(港)  /  刺激1995(台)</span>
                        </a>


                            <span class="playable">[可播放]</span>
                    </div>
                    <div class="bd">
                        <p class="">
                            导演: 弗兰克·德拉邦特 Frank Darabont&nbsp;&nbsp;&nbsp;主演: 蒂姆·罗宾斯 Tim Robbins /...<br>
                            1994&nbsp;/&nbsp;美国&nbsp;/&nbsp;犯罪 剧情
                        </p>


                        <div class="star">
                                <span class="rating5-t"></span>
                                <span class="rating_num" property="v:average">9.7</span>
                                <span property="v:best" content="10.0"></span>
                                <span>2396288人评价</span>
                        </div>

                            <p class="quote">
                                <span class="inq">希望让人自由。</span>
                            </p>
                    </div>
                </div>
            </div>
        </li>
'''