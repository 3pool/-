# -*- coding:utf-8 -*-
# 1.载入库，注意读和写是两个不同的对象
from openpyxl import load_workbook  #  读工作簿的模块
from openpyxl import Workbook   #  写入工作簿的模块
# 2.数据读取
# 2.1 工作簿的载入
myBook = load_workbook('某公司贸易数据.xlsx')
# 读取工作簿
print(myBook)  # 调试用
# 2.2 工作表的获取
mySheet = myBook.worksheets[0]  #  返回索引位置为0的表
print(mySheet)  # 调试用
# 2.3 获取工作表中的数据
print(mySheet.cell(1,1).value)  #  读取指定位置单元格的值，注意cell的索引是从1开始
print(mySheet.max_row)  #  返回数据部分的行数
print(mySheet.max_column)  # 返回数据部分的列数
print('-'*100)
# 2.4 通过嵌套循环遍历单元格，并获取数据
myList = []
# 构建一个列表，用来存储读取出来的结构化数据
for iRow in range(mySheet.max_row):
    # 外层循环处理行
    myListin = []
    # 内层的列表，用来存储每行的数据
    for iCol in range(mySheet.max_column):
        # 内层循环处理列
        print(mySheet.cell(iRow+1, iCol+1).value,end='')
        # 输出行中每列的内容，因为openpyxl的索引是从1开始的，所以iRow,iCol都要加1
        myListin.append(mySheet.cell(iRow+1, iCol+1).value)
        # 将每列的内容添加到内层列表中

    print()
    # 外层循环中换行
    myList.append(myListin)
    # 在外层循环中，将每行数据构成的列表添加到myList中，形成一个嵌套的列表，即结构化数据
print(myList)
print('-'*100)
# 3.写入数据
# 3.1 创建用来写入的工作簿
newBook = Workbook()
# 创建工作簿
# 3.2 创建工作表
newSheet = newBook.create_sheet('数据写入',0)
# 在工作簿中索引为0的位置上创建一张新的工作表
# 3.3 写入数据
# 3.3.1 写入表头
# newSheet.cell(row_index,col_index).value = 表达式
newSheet.cell(1, 1).value = '类别编号'
newSheet.cell(1, 2).value = '类别名称'
newSheet.cell(1, 3).value = '说明'
# 3.3.2 写入内容：利用嵌套的循环写入结构化的数据
for iRowx in range(len(myList)-1):
    # 外层循环处理行，myList中的 元素数量就是要写入的行数，因为表头已经单独写入，所以-1跳过表头
    for iColx in range(len(myList[iRowx+1])):
        # 内层循环处理列，列的数量是myList中每个小列表的元素数量
        newSheet.cell(iRowx + 2, iColx + 1).value = myList[iRowx+1][iColx]
        # 在指定的单元格中写入指定的值
        # iRow+2是为了跳过表头，另外，python的索引是0，cell的索引是1，所以是iRow+2,iColx+1同理
        # myList[iRowx+1][iColx]，是为了提取指定的数据，+1是为了跳过表头
        # 写入操作的核心，仍然是定位
# 3.4 把内存的工作簿写入磁盘文件
newBook.save('openpyxl写入.xlsx')